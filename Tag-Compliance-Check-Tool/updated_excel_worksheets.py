"""
Enhanced Excel Worksheets Module

Updated to use proper data types, number formatting, and enhanced styling.
"""

import re
import logging
from typing import List, Set, Optional, Dict, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config_manager import Config
from data_models import TagData, ResourceGroupTagData, SubscriptionInfo
from excel_enhanced_formatting import ExcelFormattingManager, EnhancedWorksheetFormatter
from constants import TagComplianceStatus
from utils import sanitize_for_excel

logger = logging.getLogger(__name__)


class EnhancedWorksheetGenerator:
    """Enhanced worksheet generator with proper formatting and data types"""
    
    def __init__(self, config: Config):
        self.config = config
        self.formatting_manager = ExcelFormattingManager()
        self.formatter = EnhancedWorksheetFormatter(self.formatting_manager)
    
    def create_enhanced_worksheet_with_table(self, wb: Workbook, title: str, headers: List[str], 
                                           data: List[List[Any]], 
                                           sheet_type: str = 'default') -> None:
        """Create a worksheet with enhanced formatting based on sheet type"""
        ws = wb.create_sheet(title=title)
        
        if not data:
            # Create empty sheet with headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=sanitize_for_excel(header))
                cell.style = self.formatting_manager.header_style
            return
        
        # Convert data to proper format with sanitization
        formatted_data = []
        for row in data:
            formatted_row = []
            for item in row:
                if item is None:
                    formatted_row.append("")
                elif isinstance(item, str):
                    formatted_row.append(sanitize_for_excel(item))
                else:
                    formatted_row.append(item)
            formatted_data.append(formatted_row)
        
        # Apply specific formatting based on sheet type
        if sheet_type == 'summary':
            self.formatter.format_summary_sheet(ws, formatted_data, headers)
        elif sheet_type == 'subscription':
            self.formatter.format_subscription_sheet(ws, formatted_data, headers)
        elif sheet_type == 'compliance':
            self.formatter.format_compliance_sheet(ws, formatted_data, headers)
        elif sheet_type == 'detailed':
            self.formatter.format_detailed_sheet(ws, formatted_data, headers)
        else:
            self._apply_default_formatting(ws, formatted_data, headers)
    
    def _apply_default_formatting(self, worksheet, data: List[List[Any]], headers: List[str]):
        """Apply default formatting to worksheet"""
        # Define default data types
        data_types = {}
        for i in range(len(headers)):
            data_types[i + 1] = 'text'  # Default to text
        
        self.formatter._apply_standard_formatting(worksheet, data, headers, data_types)
    
    def generate_detailed_sheets(self, wb: Workbook, resource_tags: List[TagData], 
                               rg_tags: List[ResourceGroupTagData], 
                               subscriptions: List[SubscriptionInfo]) -> None:
        """Generate detailed data sheets with enhanced formatting"""
        
        # Resource Tag Data
        tag_headers = ["Subscription", "Tag Name", "Tag Value", "Resource Name", "Resource Type", 
                      "Resource ID", "Location", "Compliance Status", "Canonical Tag", "Variation Matched"]
        tag_data = []
        
        for tag in resource_tags:
            tag_data.append([
                tag.subscription_name, tag.name, tag.value, tag.resource_name, 
                tag.resource_type, tag.resource_id, tag.resource_location,
                tag.compliance_status, tag.canonical_tag_name, tag.variation_matched
            ])
        
        self.create_enhanced_worksheet_with_table(
            wb, "Resource Tag Details", tag_headers, tag_data, 'detailed'
        )
        
        # Resource Group Tag Data
        rg_headers = ["Subscription", "RG Tag Name", "RG Tag Value", "Resource Group Name", 
                     "Resource Group ID", "Compliance Status", "Canonical Tag", "Variation Matched"]
        rg_data = []
        
        for tag in rg_tags:
            rg_data.append([
                tag.subscription_name, tag.name, tag.value, tag.resource_name, 
                tag.resource_id, tag.compliance_status, tag.canonical_tag_name, tag.variation_matched
            ])
        
        self.create_enhanced_worksheet_with_table(
            wb, "Resource Group Tag Details", rg_headers, rg_data, 'detailed'
        )
        
        # Tag values summary
        self._generate_tag_value_summary(wb, resource_tags)
    
    def _generate_tag_value_summary(self, wb: Workbook, resource_tags: List[TagData]):
        """Generate tag values summary with proper number formatting"""
        tag_value_counts = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing:
                key = (tag.canonical_tag_name or tag.name, tag.value)
                tag_value_counts[key] = tag_value_counts.get(key, 0) + 1

        tag_value_summary = []
        for (name, value), count in sorted(tag_value_counts.items(), key=lambda x: x[1], reverse=True):
            tag_value_summary.append([name, value, count])

        headers = ["Tag Key", "Tag Value", "Usage Count"]
        
        # Create worksheet
        ws = wb.create_sheet(title="Tag Values Summary")
        
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = self.formatting_manager.header_style
        
        # Apply data with proper number formatting for count column
        for row_idx, row_data in enumerate(tag_value_summary, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if col_idx == 3:  # Count column
                    cell.style = self.formatting_manager.number_style
                else:
                    cell.style = self.formatting_manager.data_style
        
        # Create table
        if tag_value_summary:
            table_range = f"A1:C{len(tag_value_summary) + 1}"
            table = Table(displayName="TagValuesSummary", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        self.formatting_manager.auto_adjust_columns(ws)
        ws.freeze_panes = 'A2'
    
    def generate_tag_summaries(self, wb: Workbook, resource_tags: List[TagData], 
                             rg_tags: List[ResourceGroupTagData]) -> None:
        """Generate tag summary sheets with proper number formatting"""
        
        # Canonical tag summary
        canonical_counts = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing and tag.canonical_tag_name:
                canonical_counts[tag.canonical_tag_name] = canonical_counts.get(tag.canonical_tag_name, 0) + 1
        
        canonical_summary = [[name, count] for name, count in sorted(canonical_counts.items(), key=lambda x: x[1], reverse=True)]
        
        self._create_summary_sheet(wb, "Canonical Tag Summary", 
                                   ["Canonical Tag Name", "Usage Count"], canonical_summary)
        
        # Raw tag names summary
        tag_counts = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing:
                tag_counts[tag.name] = tag_counts.get(tag.name, 0) + 1
        
        tag_summary = [[name, count] for name, count in sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)]
        
        self._create_summary_sheet(wb, "Raw Tag Names Summary", 
                                   ["Tag Name", "Usage Count"], tag_summary)
    
    def _create_summary_sheet(self, wb: Workbook, title: str, headers: List[str], data: List[List[Any]]):
        """Create a summary sheet with proper number formatting"""
        ws = wb.create_sheet(title=title)
        
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = self.formatting_manager.header_style
        
        # Apply data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if col_idx == 2:  # Count column
                    cell.style = self.formatting_manager.number_style
                else:
                    cell.style = self.formatting_manager.data_style
        
        # Create table
        if data:
            table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
            clean_title = re.sub(r'[^\w\s]', '', title).replace(' ', '_').strip('_')
            table = Table(displayName=f"Table_{clean_title}", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        self.formatting_manager.auto_adjust_columns(ws)
        ws.freeze_panes = 'A2'