"""
Excel Styles Module

Handles all Excel styling, fonts, colors, and formatting for the Azure Tagging Analysis Tool.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from constants import ColorThresholds


class ExcelStyleManager:
    """Manages all Excel styles and formatting"""
    
    def __init__(self):
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel styles with softer color coding for better readability"""
        # Fonts
        self.header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
        self.data_font = Font(name='Arial', color='2C3E50', size=10)
        self.bold_font = Font(name='Arial', bold=True, color='2C3E50', size=10)
        
        # Softer fills for performance levels
        self.excellent_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
        self.good_fill = PatternFill(start_color='E1F5FE', end_color='E1F5FE', fill_type='solid')
        self.fair_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        self.poor_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
        self.header_fill = PatternFill(start_color='5D87A1', end_color='5D87A1', fill_type='solid')
        self.highlight_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        self.partial_fill = PatternFill(start_color='F3E5AB', end_color='F3E5AB', fill_type='solid')
        
        # Alignments
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Borders
        thin_border = Side(border_style="thin", color="CCCCCC")
        self.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
    
    def get_percentage_fill(self, percentage: float) -> PatternFill:
        """Get appropriate fill color based on percentage"""
        if percentage >= ColorThresholds.EXCELLENT:
            return self.excellent_fill
        elif percentage >= ColorThresholds.GOOD:
            return self.good_fill
        elif percentage >= ColorThresholds.FAIR:
            return self.fair_fill
        else:
            return self.poor_fill
    
    def get_status_text(self, percentage: float) -> str:
        """Get status text based on percentage"""
        if percentage >= ColorThresholds.EXCELLENT:
            return "Excellent"
        elif percentage >= ColorThresholds.GOOD:
            return "Good"
        elif percentage >= ColorThresholds.FAIR:
            return "Needs Improvement"
        else:
            return "Poor"
    
    def get_compliance_font(self, status: str) -> Font:
        """Get appropriate font for compliance status"""
        from constants import TagComplianceStatus
        
        if status == TagComplianceStatus.COMPLIANT:
            return Font(name='Arial', bold=True, color='2E7D32', size=10)
        elif status == TagComplianceStatus.PARTIAL:
            return Font(name='Arial', bold=True, color='F57C00', size=10)
        elif status == TagComplianceStatus.NON_COMPLIANT:
            return Font(name='Arial', bold=True, color='C62828', size=10)
        else:
            return self.data_font
    
    def get_compliance_fill(self, status: str) -> PatternFill:
        """Get appropriate fill for compliance status"""
        from constants import TagComplianceStatus
        
        if status == TagComplianceStatus.COMPLIANT:
            return self.good_fill
        elif status == TagComplianceStatus.PARTIAL:
            return self.partial_fill
        elif status == TagComplianceStatus.NON_COMPLIANT:
            return self.poor_fill
        else:
            return PatternFill()
