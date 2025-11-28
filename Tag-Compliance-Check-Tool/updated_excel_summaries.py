"""
Final Fixed Excel Summaries Module
Removed title row insertion that was causing duplicate headers and Excel corruption.
"""

import logging
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from config_manager import Config
from data_models import TagData, ResourceGroupTagData, SubscriptionInfo
from excel_enhanced_formatting import ExcelFormattingManager, EnhancedWorksheetFormatter
from updated_excel_worksheets import EnhancedWorksheetGenerator
from constants import TagComplianceStatus

logger = logging.getLogger(__name__)

class EnhancedSummaryGenerator:
    """Enhanced summary generator with proper formatting and data types"""

    def __init__(self, config: Config):
        self.config = config
        self.formatting_manager = ExcelFormattingManager()
        self.formatter = EnhancedWorksheetFormatter(self.formatting_manager)
        self.worksheet_generator = EnhancedWorksheetGenerator(config)

    def generate_enhanced_summary(self, wb: Workbook, resource_tags: List[TagData],
                                rg_tags: List[ResourceGroupTagData],
                                subscriptions: List[SubscriptionInfo]) -> None:
        """Generate enhanced summary with proper number and percentage formatting"""
        
        # Calculate metrics
        total_resources = sum(sub.resource_count for sub in subscriptions)
        total_rgs = sum(sub.rg_count for sub in subscriptions)
        total_tagged_resources = sum(sub.tagged_resources for sub in subscriptions)
        total_tagged_rgs = sum(sub.tagged_rgs for sub in subscriptions)
        total_compliant_resources = sum(sub.mandatory_compliant_resources for sub in subscriptions)
        total_partial_resources = sum(sub.mandatory_partial_resources for sub in subscriptions)

        # Calculate percentages as decimal values for proper Excel percentage formatting
        resource_tagging_pct = (total_tagged_resources / total_resources) if total_resources > 0 else 0
        rg_tagging_pct = (total_tagged_rgs / total_rgs) if total_rgs > 0 else 0
        compliance_pct = (total_compliant_resources / total_resources) if total_resources > 0 else 0
        partial_pct = (total_partial_resources / total_resources) if total_resources > 0 else 0
        combined_pct = ((total_compliant_resources + total_partial_resources) / total_resources) if total_resources > 0 else 0

        missing_count = sum(1 for tag in resource_tags if tag.is_mandatory_missing)
        untagged_count = sum(1 for tag in resource_tags if tag.is_untagged)

        headers = ["Metric", "Count", "Total", "Percentage", "Status"]

        summary_data = [
            ["Resources with Tags", total_tagged_resources, total_resources, resource_tagging_pct, self._get_status_text(resource_tagging_pct * 100)],
            ["Resource Groups with Tags", total_tagged_rgs, total_rgs, rg_tagging_pct, self._get_status_text(rg_tagging_pct * 100)],
            ["Full Mandatory Compliance", total_compliant_resources, total_resources, compliance_pct, self._get_status_text(compliance_pct * 100)],
            ["Partial Compliance (Variations)", total_partial_resources, total_resources, partial_pct, self._get_status_text(partial_pct * 100)],
            ["Combined Compliance", total_compliant_resources + total_partial_resources, total_resources, combined_pct, self._get_status_text(combined_pct * 100)],
            ["", "", "", "", ""],  # Separator row
            ["Total Subscriptions", len(subscriptions), len(subscriptions), 1.0, "Complete"],
            ["Total Resources", total_resources, total_resources, 1.0, "Complete"],
            ["Total Resource Groups", total_rgs, total_rgs, 1.0, "Complete"],
            ["Untagged Resources", untagged_count, total_resources, (untagged_count/total_resources) if total_resources > 0 else 0, ""],
            ["Missing Mandatory Tags", missing_count, "", "", ""],
            ["Mandatory Tags Defined", len([t for t in self.config.mandatory_tags if t != "NONE"]), "", "", ""],
            ["Tag Variations Configured", len(self.config.tag_variations), "", "", ""],
        ]

        # Create worksheet manually without title row to avoid corruption
        ws = wb.create_sheet(title="Executive Summary")
        
        # Apply formatting using the formatter
        self.formatter.format_summary_sheet(ws, summary_data, headers)
        
        logger.debug("✅ Executive Summary generated successfully")

    def generate_subscription_analysis(self, wb: Workbook, subscriptions: List[SubscriptionInfo]) -> None:
        """Generate subscription-level analysis with proper percentage formatting"""
        
        headers = [
            "Subscription Name", "Subscription ID", "Resources", "Tagged Resources",
            "Resource Tag %", "Resource Groups", "Tagged RGs", "RG Tag %",
            "Full Compliant", "Full Compliance %", "Partial Compliant", "Partial %", "Combined %"
        ]

        sub_data = []
        for sub in sorted(subscriptions, key=lambda x: x.combined_compliance_percentage, reverse=True):
            sub_data.append([
                sub.name,
                sub.id,
                sub.resource_count,
                sub.tagged_resources,
                sub.resource_tagging_percentage / 100,  # Convert to decimal for Excel percentage format
                sub.rg_count,
                sub.tagged_rgs,
                sub.rg_tagging_percentage / 100,  # Convert to decimal
                sub.mandatory_compliant_resources,
                sub.mandatory_compliance_percentage / 100,  # Convert to decimal
                sub.mandatory_partial_resources,
                sub.mandatory_partial_percentage / 100,  # Convert to decimal
                sub.combined_compliance_percentage / 100  # Convert to decimal
            ])

        # Create worksheet manually
        ws = wb.create_sheet(title="Subscription Analysis")
        self.formatter.format_subscription_sheet(ws, sub_data, headers)
        
        logger.debug("✅ Subscription Analysis generated successfully")

    def generate_compliance_report(self, wb: Workbook, resource_tags: List[TagData],
                                 subscriptions: List[SubscriptionInfo]) -> None:
        """Generate mandatory tag compliance report with proper formatting"""
        
        if not self.config.mandatory_tags or self.config.mandatory_tags == ["NONE"]:
            logger.debug("⏭️ Skipping compliance report - no mandatory tags defined")
            return

        compliance_data = {}
        for sub in subscriptions:
            sub_resources = [tag for tag in resource_tags if tag.subscription_name == sub.name]
            
            for mandatory_tag in self.config.mandatory_tags:
                if mandatory_tag == "NONE":
                    continue

                full_compliant = len(set(
                    tag.resource_id for tag in sub_resources
                    if tag.canonical_tag_name == mandatory_tag and tag.compliance_status == TagComplianceStatus.COMPLIANT
                ))

                partial_compliant = len(set(
                    tag.resource_id for tag in sub_resources
                    if tag.canonical_tag_name == mandatory_tag and tag.compliance_status == TagComplianceStatus.PARTIAL
                ))

                non_compliant = sub.resource_count - full_compliant - partial_compliant

                # Store as decimal values for Excel percentage formatting
                full_pct = (full_compliant / sub.resource_count) if sub.resource_count > 0 else 0
                partial_pct = (partial_compliant / sub.resource_count) if sub.resource_count > 0 else 0
                combined_pct = ((full_compliant + partial_compliant) / sub.resource_count) if sub.resource_count > 0 else 0

                compliance_data[(sub.name, mandatory_tag)] = {
                    'subscription': sub.name, 'tag': mandatory_tag,
                    'full_compliant': full_compliant, 'partial_compliant': partial_compliant,
                    'non_compliant': non_compliant, 'total_resources': sub.resource_count,
                    'full_percentage': full_pct, 'partial_percentage': partial_pct,
                    'combined_percentage': combined_pct
                }

        headers = ["Subscription", "Mandatory Tag", "Full Compliant", "Full %", "Partial Compliant",
                  "Partial %", "Non-Compliant", "Total Resources", "Combined %"]

        compliance_report_data = []
        for key, data in sorted(compliance_data.items()):
            compliance_report_data.append([
                data['subscription'],
                data['tag'],
                data['full_compliant'],
                data['full_percentage'],  # Decimal value for Excel percentage
                data['partial_compliant'],
                data['partial_percentage'],  # Decimal value for Excel percentage
                data['non_compliant'],
                data['total_resources'],
                data['combined_percentage']  # Decimal value for Excel percentage
            ])

        # Create worksheet manually
        ws = wb.create_sheet(title="Compliance Report")
        self.formatter.format_compliance_sheet(ws, compliance_report_data, headers)
        
        logger.debug("✅ Compliance Report generated successfully")

    def generate_tag_variation_analysis(self, wb: Workbook, resource_tags: List[TagData],
                                      rg_tags: List[ResourceGroupTagData]) -> None:
        """Generate tag variation analysis report with proper formatting"""
        
        if not self.config.tag_variations:
            logger.debug("⏭️ Skipping tag variation analysis - no variations defined")
            return

        variation_data = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing and not tag.is_untagged:
                canonical = tag.canonical_tag_name
                if canonical in self.config.tag_variations:
                    if canonical not in variation_data:
                        variation_data[canonical] = {
                            'canonical': canonical, 'variations': {}, 'total_usage': 0
                        }

                    variation_key = tag.name
                    if variation_key not in variation_data[canonical]['variations']:
                        variation_data[canonical]['variations'][variation_key] = {
                            'count': 0, 'status': tag.compliance_status, 'examples': set()
                        }

                    variation_data[canonical]['variations'][variation_key]['count'] += 1
                    variation_data[canonical]['variations'][variation_key]['examples'].add(f"{tag.resource_name} ({tag.subscription_name})")
                    variation_data[canonical]['total_usage'] += 1

        headers = ["Canonical Tag", "Variation Used", "Usage Count", "Usage %", "Compliance Status", "Example Resources"]

        variation_report_data = []
        for canonical, data in sorted(variation_data.items()):
            for variation, var_data in sorted(data['variations'].items(), key=lambda x: x[1]['count'], reverse=True):
                # Calculate percentage as decimal for Excel formatting
                usage_pct = (var_data['count'] / data['total_usage']) if data['total_usage'] > 0 else 0

                examples = "; ".join(list(var_data['examples'])[:3])
                if len(var_data['examples']) > 3:
                    examples += f" ... (+{len(var_data['examples']) - 3} more)"

                variation_report_data.append([
                    canonical,
                    variation,
                    var_data['count'],
                    usage_pct,  # Decimal value for Excel percentage
                    var_data['status'],
                    examples
                ])

        # Create worksheet manually with careful formatting
        ws = wb.create_sheet(title="Tag Variations Analysis")

        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.formatting_manager._create_border()

        # Apply data with specific formatting
        for row_idx, row_data in enumerate(variation_report_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if col_idx == 3:  # Usage Count
                    cell.number_format = '#,##0'
                    cell.alignment = self.formatting_manager.number_style.alignment
                elif col_idx == 4:  # Usage %
                    cell.number_format = self.formatting_manager.formats['percentage']
                    cell.alignment = self.formatting_manager.percentage_style.alignment
                    cell.font = Font(name='Segoe UI', color='2C3E50', size=10, bold=True)
                elif col_idx == 5:  # Compliance Status
                    self._apply_compliance_cell_formatting(cell, value)
                else:
                    cell.alignment = self.formatting_manager.data_style.alignment
                    cell.font = Font(name='Segoe UI', color='2C3E50', size=10)

                cell.border = self.formatting_manager._create_border()

        # Create table safely
        if variation_report_data:
            try:
                table_range = f"A1:F{len(variation_report_data) + 1}"
                table = Table(displayName="TagVariationsAnalysis", ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            except Exception as e:
                logger.warning(f"Could not create table for Tag Variations Analysis: {e}")

        self.formatting_manager.auto_adjust_columns(ws)
        ws.freeze_panes = 'A2'
        
        logger.debug("✅ Tag Variations Analysis generated successfully")

    def _get_status_text(self, percentage: float) -> str:
        """Get status text based on percentage"""
        from constants import ColorThresholds
        
        if percentage >= ColorThresholds.EXCELLENT:
            return "Excellent"
        elif percentage >= ColorThresholds.GOOD:
            return "Good"
        elif percentage >= ColorThresholds.FAIR:
            return "Needs Improvement"
        else:
            return "Poor"

    def _apply_compliance_cell_formatting(self, cell, value: str):
        """Apply compliance-specific cell formatting"""
        value_lower = str(value).lower() if value else ""
        
        if 'compliant' in value_lower and 'non' not in value_lower:
            colors = self.formatting_manager.compliance_colors['compliant']
        elif 'partial' in value_lower:
            colors = self.formatting_manager.compliance_colors['partial']
        else:
            colors = self.formatting_manager.compliance_colors['non_compliant']
        
        cell.fill = PatternFill(start_color=colors['fill'], 
                               end_color=colors['fill'], fill_type='solid')
        cell.font = Font(name='Segoe UI', color=colors['font'], bold=True, size=10)
