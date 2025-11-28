"""
Excel Generator - Core Module

Main orchestrator for Excel report generation.
"""

import logging
from typing import List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config_manager import Config
from data_models import TagData, ResourceGroupTagData, SubscriptionInfo
from excel_styles import ExcelStyleManager
from excel_worksheets import WorksheetGenerator
from excel_summaries import SummaryGenerator

logger = logging.getLogger(__name__)


class EnhancedExcelReportGenerator:
    """Generate enhanced Excel reports with tables, color coding, and improved formatting"""
    
    def __init__(self, config: Config):
        self.config = config
        self.wb = Workbook()
        self.style_manager = ExcelStyleManager()
        self.worksheet_generator = WorksheetGenerator(config, self.style_manager)
        self.summary_generator = SummaryGenerator(config, self.style_manager)
    
    def generate_report(self, resource_tags: List[TagData], rg_tags: List[ResourceGroupTagData], 
                       subscriptions: List[SubscriptionInfo]) -> None:
        """Generate complete enhanced Excel report"""
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Generate all sheets using sub-modules
        self.summary_generator.generate_enhanced_summary(
            self.wb, resource_tags, rg_tags, subscriptions
        )
        self.summary_generator.generate_subscription_analysis(
            self.wb, subscriptions
        )
        self.summary_generator.generate_compliance_report(
            self.wb, resource_tags, subscriptions
        )
        self.summary_generator.generate_tag_variation_analysis(
            self.wb, resource_tags, rg_tags
        )
        
        # Generate detailed sheets
        self.worksheet_generator.generate_detailed_sheets(
            self.wb, resource_tags, rg_tags, subscriptions
        )
        self.worksheet_generator.generate_tag_summaries(
            self.wb, resource_tags, rg_tags
        )
    
    def save(self, filename: str) -> None:
        """Save the workbook with enhanced formatting"""
        
        if "Executive Summary" in self.wb.sheetnames:
            self.wb.active = self.wb["Executive Summary"]
        
        self.wb.save(filename)
        logger.info(f"Enhanced report with tables saved to {filename}")
