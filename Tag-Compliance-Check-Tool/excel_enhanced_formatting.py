"""
Fixed Excel Enhanced Formatting Module

Corrected header application and percentage formatting issues.
"""

import logging
import re
from typing import List, Any, Dict, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.layout import Layout, ManualLayout

from constants import ColorThresholds
from utils import sanitize_for_excel

logger = logging.getLogger(__name__)


class ExcelFormattingManager:
    """Advanced Excel formatting manager with proper data types and styling"""
    
    def __init__(self):
        self._setup_named_styles()
        self._setup_number_formats()
        self._setup_colors()
    
    def _setup_named_styles(self):
        """Setup named styles for consistent formatting"""
        # Header style
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
        self.header_style.fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.header_style.border = self._create_border()
        
        # Data style
        self.data_style = NamedStyle(name="data_style")
        self.data_style.font = Font(name='Segoe UI', color='2C3E50', size=10)
        self.data_style.alignment = Alignment(vertical='center')
        self.data_style.border = self._create_border()
        
        # Percentage style
        self.percentage_style = NamedStyle(name="percentage_style")
        self.percentage_style.font = Font(name='Segoe UI', color='2C3E50', size=10, bold=True)
        self.percentage_style.alignment = Alignment(horizontal='center', vertical='center')
        self.percentage_style.number_format = FORMAT_PERCENTAGE_00
        self.percentage_style.border = self._create_border()
        
        # Number style
        self.number_style = NamedStyle(name="number_style")
        self.number_style.font = Font(name='Segoe UI', color='2C3E50', size=10)
        self.number_style.alignment = Alignment(horizontal='right', vertical='center')
        self.number_style.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        self.number_style.border = self._create_border()
        
        # Currency style
        self.currency_style = NamedStyle(name="currency_style")
        self.currency_style.font = Font(name='Segoe UI', color='2C3E50', size=10)
        self.currency_style.alignment = Alignment(horizontal='right', vertical='center')
        self.currency_style.number_format = '"$"#,##0.00'
        self.currency_style.border = self._create_border()
        
        # Date style
        self.date_style = NamedStyle(name="date_style")
        self.date_style.font = Font(name='Segoe UI', color='2C3E50', size=10)
        self.date_style.alignment = Alignment(horizontal='center', vertical='center')
        self.date_style.number_format = 'MM/DD/YYYY'
        self.date_style.border = self._create_border()
    
    def _setup_number_formats(self):
        """Setup custom number formats"""
        self.formats = {
            'percentage': FORMAT_PERCENTAGE_00,
            'number': FORMAT_NUMBER_COMMA_SEPARATED1,
            'currency': '"$"#,##0.00',
            'date': 'MM/DD/YYYY',
            'datetime': 'MM/DD/YYYY HH:MM',
            'integer': '#,##0',
            'decimal_2': '#,##0.00',
            'text': '@'
        }
    
    def _setup_colors(self):
        """Setup color schemes for different performance levels"""
        self.performance_colors = {
            'excellent': {'fill': 'C6EFCE', 'font': '006100'},  # Light green background, dark green text
            'good': {'fill': 'D4F4FF', 'font': '0066CC'},       # Light blue background, dark blue text
            'warning': {'fill': 'FFEB9C', 'font': 'CC6600'},    # Light yellow background, dark orange text
            'poor': {'fill': 'FFC7CE', 'font': 'CC0000'}        # Light red background, dark red text
        }
        
        self.compliance_colors = {
            'compliant': {'fill': 'C6EFCE', 'font': '006100'},
            'partial': {'fill': 'FFEB9C', 'font': 'CC6600'},
            'non_compliant': {'fill': 'FFC7CE', 'font': 'CC0000'}
        }
    
    def _create_border(self):
        """Create consistent border style"""
        thin_border = Side(border_style="thin", color="B0B0B0")
        return Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
    
    def apply_conditional_formatting(self, worksheet, col_idx: int, start_row: int, end_row: int, 
                                   format_type: str = 'percentage'):
        """Apply conditional formatting based on performance thresholds"""
        column_letter = get_column_letter(col_idx)
        
        for row in range(start_row, end_row + 1):
            cell = worksheet[f"{column_letter}{row}"]
            
            try:
                if format_type == 'percentage':
                    # Handle percentage values - they should already be decimal values
                    value = cell.value
                    if isinstance(value, (int, float)):
                        numeric_value = float(value) * 100  # Convert decimal to percentage for color logic
                    else:
                        continue
                else:
                    numeric_value = float(cell.value) if cell.value is not None else 0
                
                # Apply color based on performance
                color_scheme = self._get_performance_color_scheme(numeric_value)
                cell.fill = PatternFill(start_color=color_scheme['fill'], 
                                      end_color=color_scheme['fill'], fill_type='solid')
                cell.font = Font(name='Segoe UI', color=color_scheme['font'], bold=True, size=10)
                
            except (ValueError, TypeError):
                continue
    
    def _get_performance_color_scheme(self, value: float) -> Dict[str, str]:
        """Get color scheme based on performance value"""
        if value >= ColorThresholds.EXCELLENT:
            return self.performance_colors['excellent']
        elif value >= ColorThresholds.GOOD:
            return self.performance_colors['good']
        elif value >= ColorThresholds.FAIR:
            return self.performance_colors['warning']
        else:
            return self.performance_colors['poor']
    
    def create_enhanced_table(self, worksheet, table_name: str, data_range: str, 
                            data_types: Optional[Dict[int, str]] = None):
        """Create an enhanced Excel table with proper formatting"""
        # Create table
        table = Table(displayName=table_name, ref=data_range)
        
        # Apply table style
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)
    
    def add_data_bars(self, worksheet, col_idx: int, start_row: int, end_row: int):
        """Add data bars for visual representation"""
        column_letter = get_column_letter(col_idx)
        range_address = f"{column_letter}{start_row}:{column_letter}{end_row}"
        
        data_bar = DataBarRule(
            start_type='num', start_value=0,
            end_type='num', end_value=1,  # For percentage values (0-1)
            color="4472C4"
        )
        worksheet.conditional_formatting.add(range_address, data_bar)
    
    def auto_adjust_columns(self, worksheet, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width


class EnhancedWorksheetFormatter:
    """Enhanced worksheet formatter with proper data types and styling"""
    
    def __init__(self, formatting_manager: ExcelFormattingManager):
        self.formatter = formatting_manager
    
    def format_summary_sheet(self, worksheet, data: List[List[Any]], headers: List[str]):
        """Format executive summary sheet with proper data types"""
        # Apply header formatting FIRST
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.formatter._create_border()
        
        # Apply data formatting
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Handle different column types
                if col_idx == 1:  # Metric name
                    cell.value = str(value) if value is not None else ""
                    cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col_idx == 2:  # Count
                    cell.value = int(value) if isinstance(value, (int, float)) and value != "" else value
                    cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0'
                elif col_idx == 3:  # Total
                    cell.value = int(value) if isinstance(value, (int, float)) and value != "" else value
                    cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0'
                elif col_idx == 4:  # Percentage
                    if isinstance(value, (int, float)) and value != "":
                        cell.value = float(value)  # Should already be decimal (0.0-1.0)
                        cell.number_format = FORMAT_PERCENTAGE_00
                        cell.font = Font(name='Segoe UI', color='2C3E50', size=10, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.value = value
                        cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 5:  # Status
                    cell.value = str(value) if value is not None else ""
                    cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply border to all cells
                cell.border = self.formatter._create_border()
        
        # Apply conditional formatting to percentage column
        if len(data) > 0:
            self.formatter.apply_conditional_formatting(
                worksheet, 4, 2, len(data) + 1, 'percentage'
            )
        
        # Create table
        if len(data) > 0:
            table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
            self.formatter.create_enhanced_table(worksheet, "SummaryTable", table_range)
        
        # Auto-adjust columns
        self.formatter.auto_adjust_columns(worksheet)
        
        # Freeze panes
        worksheet.freeze_panes = 'A2'
    
    def format_subscription_sheet(self, worksheet, data: List[List[Any]], headers: List[str]):
        """Format subscription analysis sheet"""
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.formatter._create_border()
        
        # Apply data with proper formatting
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Format based on column type
                if col_idx in [3, 4, 6, 7, 9, 11]:  # Number columns
                    cell.value = int(value) if isinstance(value, (int, float)) else value
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx in [5, 8, 10, 12, 13]:  # Percentage columns
                    if isinstance(value, (int, float)):
                        cell.value = float(value)
                        cell.number_format = FORMAT_PERCENTAGE_00
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(name='Segoe UI', color='2C3E50', size=10, bold=True)
                    else:
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                else:  # Text columns
                    cell.value = str(value) if value is not None else ""
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                cell.border = self.formatter._create_border()
        
        # Apply conditional formatting to percentage columns
        percentage_columns = [5, 8, 10, 12, 13]
        for col in percentage_columns:
            if len(data) > 0:
                self.formatter.apply_conditional_formatting(
                    worksheet, col, 2, len(data) + 1, 'percentage'
                )
        
        # Add data bars to combined compliance column
        if len(data) > 0:
            self.formatter.add_data_bars(worksheet, 13, 2, len(data) + 1)
        
        # Create table and format
        if len(data) > 0:
            table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
            self.formatter.create_enhanced_table(worksheet, "SubscriptionTable", table_range)
        
        self.formatter.auto_adjust_columns(worksheet)
        worksheet.freeze_panes = 'A2'
    
    def format_compliance_sheet(self, worksheet, data: List[List[Any]], headers: List[str]):
        """Format compliance report sheet"""
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.formatter._create_border()
        
        # Apply data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if col_idx in [3, 5, 7, 8]:  # Number columns
                    cell.value = int(value) if isinstance(value, (int, float)) else value
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx in [4, 6, 9]:  # Percentage columns
                    if isinstance(value, (int, float)):
                        cell.value = float(value)
                        cell.number_format = FORMAT_PERCENTAGE_00
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(name='Segoe UI', color='2C3E50', size=10, bold=True)
                    else:
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                else:  # Text columns
                    cell.value = str(value) if value is not None else ""
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                cell.border = self.formatter._create_border()
        
        # Apply conditional formatting
        percentage_columns = [4, 6, 9]
        for col in percentage_columns:
            if len(data) > 0:
                self.formatter.apply_conditional_formatting(
                    worksheet, col, 2, len(data) + 1, 'percentage'
                )
        
        if len(data) > 0:
            table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
            self.formatter.create_enhanced_table(worksheet, "ComplianceTable", table_range)
        
        self.formatter.auto_adjust_columns(worksheet)
        worksheet.freeze_panes = 'A2'
    
    def format_detailed_sheet(self, worksheet, data: List[List[Any]], headers: List[str]):
        """Format detailed data sheets"""
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.formatter._create_border()
        
        # Apply data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                cell.font = Font(name='Segoe UI', color='2C3E50', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.formatter._create_border()
                
                # Apply compliance status formatting for status columns
                if col_idx == 8:  # Compliance Status column
                    self._apply_compliance_formatting(cell, value)
        
        if len(data) > 0:
            table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
            clean_title = worksheet.title.replace(' ', '_').replace('-', '_')
            self.formatter.create_enhanced_table(worksheet, f"Table_{clean_title}", table_range)
        
        self.formatter.auto_adjust_columns(worksheet)
        worksheet.freeze_panes = 'A2'
    
    def _apply_compliance_formatting(self, cell, value):
        """Apply compliance-specific formatting"""
        value_str = str(value).lower() if value else ""
        
        if 'compliant' in value_str and 'non' not in value_str:
            colors = self.formatter.compliance_colors['compliant']
        elif 'partial' in value_str:
            colors = self.formatter.compliance_colors['partial']
        else:
            colors = self.formatter.compliance_colors['non_compliant']
        
        cell.fill = PatternFill(start_color=colors['fill'], end_color=colors['fill'], fill_type='solid')
        cell.font = Font(name='Segoe UI', color=colors['font'], bold=True, size=10)