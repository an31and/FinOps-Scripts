"""
Final Fixed Excel Generator - Core Module

Corrected Excel file corruption issues and header problems.
"""

import logging
from typing import List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config_manager import Config
from data_models import TagData, ResourceGroupTagData, SubscriptionInfo
from excel_enhanced_formatting import ExcelFormattingManager
from updated_excel_worksheets import EnhancedWorksheetGenerator
from updated_excel_summaries import EnhancedSummaryGenerator

logger = logging.getLogger(__name__)


class UltraEnhancedExcelReportGenerator:
    """
    Ultra-enhanced Excel report generator with:
    - Fixed Excel file corruption issues
    - Proper header handling
    - Professional table styling
    - Conditional formatting with color coding
    """
    
    def __init__(self, config: Config):
        self.config = config
        self.wb = Workbook()
        self.formatting_manager = ExcelFormattingManager()
        self.worksheet_generator = EnhancedWorksheetGenerator(config)
        self.summary_generator = EnhancedSummaryGenerator(config)
        
        # Add named styles to workbook safely
        self._register_named_styles()
    
    def _register_named_styles(self):
        """Register named styles with the workbook safely"""
        try:
            # Check if styles already exist before adding
            existing_styles = [style.name for style in self.wb.named_styles]
            
            styles_to_add = [
                ('header_style', self.formatting_manager.header_style),
                ('data_style', self.formatting_manager.data_style),
                ('percentage_style', self.formatting_manager.percentage_style),
                ('number_style', self.formatting_manager.number_style),
                ('currency_style', self.formatting_manager.currency_style),
                ('date_style', self.formatting_manager.date_style)
            ]
            
            for style_name, style_obj in styles_to_add:
                if style_name not in existing_styles:
                    self.wb.add_named_style(style_obj)
                    
            logger.debug("Named styles registered successfully")
        except Exception as e:
            logger.warning(f"Some named styles couldn't be registered: {e}")
    
    def generate_report(self, resource_tags: List[TagData], rg_tags: List[ResourceGroupTagData], 
                       subscriptions: List[SubscriptionInfo]) -> None:
        """Generate complete ultra-enhanced Excel report"""
        
        logger.info("🔧 Generating ultra-enhanced Excel report with proper formatting...")
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        try:
            # Log resource scan information
            total_resources = sum(sub.resource_count for sub in subscriptions)
            logger.info(f"📊 Processing {total_resources:,} resources across {len(subscriptions)} subscriptions")
            logger.info(f"📋 Resource tags collected: {len(resource_tags):,}")
            logger.info(f"📁 Resource group tags collected: {len(rg_tags):,}")
            
            # Check for resource limits
            if total_resources > 50000:
                logger.warning(f"⚠️  Large dataset detected ({total_resources:,} resources). Report generation may take longer.")
            
            # Generate summary sheets with enhanced formatting
            logger.debug("📊 Generating executive summary...")
            self.summary_generator.generate_enhanced_summary(
                self.wb, resource_tags, rg_tags, subscriptions
            )
            
            logger.debug("📈 Generating subscription analysis...")
            self.summary_generator.generate_subscription_analysis(
                self.wb, subscriptions
            )
            
            logger.debug("🎯 Generating compliance report...")
            self.summary_generator.generate_compliance_report(
                self.wb, resource_tags, subscriptions
            )
            
            logger.debug("🔄 Generating tag variation analysis...")
            self.summary_generator.generate_tag_variation_analysis(
                self.wb, resource_tags, rg_tags
            )
            
            # Generate detailed sheets with size limits
            logger.debug("📋 Generating detailed data sheets...")
            self._generate_detailed_sheets_with_limits(resource_tags, rg_tags, subscriptions)
            
            logger.debug("📊 Generating tag summaries...")
            self.worksheet_generator.generate_tag_summaries(
                self.wb, resource_tags, rg_tags
            )
            
            # Apply final touches
            self._apply_workbook_properties()
            
            logger.info("✅ Ultra-enhanced Excel report generation complete")
            
        except Exception as e:
            logger.error(f"❌ Error generating Excel report: {e}")
            raise
    
    def _generate_detailed_sheets_with_limits(self, resource_tags: List[TagData], 
                                            rg_tags: List[ResourceGroupTagData], 
                                            subscriptions: List[SubscriptionInfo]):
        """Generate detailed sheets with size limits to prevent Excel issues"""
        
        # Excel has limits - be conservative
        MAX_ROWS_PER_SHEET = 100000
        MAX_TOTAL_CELLS = 1000000
        
        resource_tag_count = len(resource_tags)
        rg_tag_count = len(rg_tags)
        
        logger.info(f"📊 Detailed data: {resource_tag_count:,} resource tags, {rg_tag_count:,} RG tags")
        
        if resource_tag_count > MAX_ROWS_PER_SHEET:
            logger.warning(f"⚠️  Large resource tag dataset ({resource_tag_count:,} rows). Consider filtering data.")
            # Truncate to prevent Excel issues
            resource_tags = resource_tags[:MAX_ROWS_PER_SHEET]
            logger.warning(f"📊 Truncated to first {MAX_ROWS_PER_SHEET:,} resource tags for Excel compatibility")
        
        if rg_tag_count > MAX_ROWS_PER_SHEET:
            logger.warning(f"⚠️  Large RG tag dataset ({rg_tag_count:,} rows). Consider filtering data.")
            rg_tags = rg_tags[:MAX_ROWS_PER_SHEET]
            logger.warning(f"📁 Truncated to first {MAX_ROWS_PER_SHEET:,} RG tags for Excel compatibility")
        
        # Generate the detailed sheets
        self.worksheet_generator.generate_detailed_sheets(
            self.wb, resource_tags, rg_tags, subscriptions
        )
    
    def _apply_workbook_properties(self):
        """Apply workbook-level properties and final formatting"""
        # Set workbook properties
        self.wb.properties.title = "Azure Tagging Analysis Report"
        self.wb.properties.subject = "Azure Resource Tagging Compliance Analysis"
        self.wb.properties.creator = "Azure Tagging Analysis Tool v2.0"
        self.wb.properties.description = "Comprehensive analysis of Azure resource tagging compliance with variation detection"
        
        # Set active sheet to Executive Summary if it exists
        if "Executive Summary" in self.wb.sheetnames:
            self.wb.active = self.wb["Executive Summary"]
            logger.debug("Set Executive Summary as active sheet")
        
        # Apply additional formatting to key sheets - REMOVED to prevent corruption
        # self._enhance_key_sheets()  # Commented out to avoid Excel corruption
    
    def save(self, filename: str) -> None:
        """Save the workbook with all enhancements"""
        try:
            # Final validation and cleanup
            self._validate_and_cleanup()
            
            # Save the workbook
            self.wb.save(filename)
            logger.info(f"📊 Ultra-enhanced Excel report saved to: {filename}")
            
            # Log report statistics
            self._log_report_statistics()
            
        except Exception as e:
            logger.error(f"❌ Error saving Excel report: {e}")
            raise
    
    def _validate_and_cleanup(self):
        """Validate and cleanup the workbook before saving"""
        # Remove any empty sheets
        sheets_to_remove = []
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if ws.max_row <= 1 and ws.max_column <= 1:
                if ws.cell(1, 1).value is None or ws.cell(1, 1).value == "":
                    sheets_to_remove.append(sheet_name)
        
        for sheet_name in sheets_to_remove:
            if len(self.wb.sheetnames) > 1:  # Don't remove the last sheet
                self.wb.remove(self.wb[sheet_name])
                logger.debug(f"Removed empty sheet: {sheet_name}")
        
        # Ensure we have at least one sheet
        if not self.wb.sheetnames:
            self.wb.create_sheet("Summary")
            logger.warning("Created fallback Summary sheet")
        
        # Validate table names to prevent Excel corruption
        self._validate_table_names()
    
    def _validate_table_names(self):
        """Validate table names to prevent Excel corruption"""
        used_names = set()
        
        for ws in self.wb.worksheets:
            for table in ws.tables.values():
                original_name = table.displayName
                # Ensure unique names
                if original_name in used_names:
                    counter = 1
                    new_name = f"{original_name}_{counter}"
                    while new_name in used_names:
                        counter += 1
                        new_name = f"{original_name}_{counter}"
                    table.displayName = new_name
                    logger.debug(f"Renamed duplicate table: {original_name} -> {new_name}")
                
                used_names.add(table.displayName)
    
    def _log_report_statistics(self):
        """Log statistics about the generated report"""
        sheet_count = len(self.wb.sheetnames)
        total_rows = sum(ws.max_row for ws in self.wb.worksheets)
        total_cells = sum(ws.max_row * ws.max_column for ws in self.wb.worksheets)
        
        logger.info(f"📈 Report Statistics:")
        logger.info(f"   📋 Sheets created: {sheet_count}")
        logger.info(f"   📊 Total rows: {total_rows:,}")
        logger.info(f"   🔢 Total cells: {total_cells:,}")
        logger.info(f"   🎨 Enhanced formatting applied: Tables, Percentages, Numbers, Colors")
        logger.info(f"   📱 Professional styling: Conditional formatting, Auto-sizing")
        
        # Resource scan limits information
        logger.info(f"📊 Resource Scan Information:")
        logger.info(f"   🔍 No hard limits on resource scanning")
        logger.info(f"   ⚡ Performance may decrease with >50,000 resources")
        logger.info(f"   📄 Excel sheets limited to ~100,000 rows for stability")
        logger.info(f"   💾 Large datasets automatically truncated for Excel compatibility")


# For backward compatibility, create an alias
class EnhancedExcelReportGenerator(UltraEnhancedExcelReportGenerator):
    """Alias for backward compatibility"""
    pass