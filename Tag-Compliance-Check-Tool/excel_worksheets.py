"""
Excel Worksheets Module

Handles detailed worksheet generation for the Azure Tagging Analysis Tool.
"""

import re
import logging
from typing import List, Set, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config_manager import Config
from data_models import TagData, ResourceGroupTagData, SubscriptionInfo
from excel_styles import ExcelStyleManager
from constants import TagComplianceStatus
from utils import sanitize_for_excel

logger = logging.getLogger(__name__)


class WorksheetGenerator:
    """Generates detailed worksheets and tables"""
    
    def __init__(self, config: Config, style_manager: ExcelStyleManager):
        self.config = config
        self.style_manager = style_manager
    
    def _calculate_column_width(self, values: List[str], header: str) -> int:
        """Calculate optimal column width based on content"""
        max_length = len(header)
        for value in values:
            if value is not None:
                value_length = len(str(value))
                if value_length > max_length:
                    max_length = value_length
        
        optimal_width = min(max(max_length + 2, 10), 150)
        return optimal_width
    
    def create_enhanced_worksheet_with_table(self, wb: Workbook, title: str, headers: List[str], 
                                           data: List[List], highlight_rows: Optional[Set[int]] = None,
                                           percentage_columns: Optional[List[int]] = None,
                                           compliance_columns: Optional[List[int]] = None) -> None:
        """Create a worksheet with Excel table formatting and enhanced features"""
        ws = wb.create_sheet(title=title)
        
        if not data:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=sanitize_for_excel(header))
                cell.font = self.style_manager.header_font
                cell.fill = self.style_manager.header_fill
                cell.alignment = self.style_manager.center_alignment
                cell.border = self.style_manager.border
            return
        
        # Calculate column widths
        column_widths = []
        for col_idx, header in enumerate(headers):
            column_data = [str(row[col_idx]) if col_idx < len(row) and row[col_idx] is not None else "" 
                          for row in data]
            width = self._calculate_column_width(column_data, header)
            column_widths.append(width)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=sanitize_for_excel(header))
            cell.font = self.style_manager.header_font
            cell.fill = self.style_manager.header_fill
            cell.alignment = self.style_manager.center_alignment
            cell.border = self.style_manager.border
        
        # Write data with conditional formatting
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sanitized_value = sanitize_for_excel(str(value)) if value is not None else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=sanitized_value)
                cell.font = self.style_manager.data_font
                cell.border = self.style_manager.border
                
                # Apply percentage-based coloring
                if percentage_columns and col_idx in percentage_columns:
                    try:
                        if isinstance(value, str) and '%' in value:
                            percentage = float(value.replace('%', ''))
                        elif isinstance(value, (int, float)):
                            percentage = float(value)
                        else:
                            percentage = 0
                        
                        cell.fill = self.style_manager.get_percentage_fill(percentage)
                        cell.font = self.style_manager.bold_font
                        cell.alignment = self.style_manager.center_alignment
                    except (ValueError, TypeError):
                        pass
                
                # Apply compliance status coloring
                elif compliance_columns and col_idx in compliance_columns:
                    cell.fill = self.style_manager.get_compliance_fill(value)
                    cell.font = self.style_manager.get_compliance_font(value)
                    cell.alignment = self.style_manager.center_alignment
                
                # Highlight specific rows
                elif highlight_rows and row_idx in highlight_rows:
                    cell.fill = self.style_manager.highlight_fill
                
                # Set alignment
                if isinstance(value, (int, float)) or (isinstance(value, str) and '%' in value):
                    cell.alignment = self.style_manager.right_alignment
                else:
                    cell.alignment = self.style_manager.left_alignment
        
        # Set column widths
        for col_idx, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = width
        
        # Create Excel table
        if data:
            table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
            clean_title = re.sub(r'[^\w\s]', '', title).replace(' ', '_').strip('_')
            table = Table(displayName=f"Table_{clean_title}", ref=table_range)
            
            style = TableStyleInfo(
                name="TableStyleMedium2", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        ws.freeze_panes = ws['A2']
    
    def generate_detailed_sheets(self, wb: Workbook, resource_tags: List[TagData], 
                               rg_tags: List[ResourceGroupTagData], 
                               subscriptions: List[SubscriptionInfo]) -> None:
        """Generate detailed data sheets"""
        
        # Resource Tag Data
        tag_headers = ["Subscription", "Tag Name", "Tag Value", "Resource Name", "Resource Type", 
                      "Resource ID", "Location", "Compliance Status", "Canonical Tag", "Variation Matched"]
        tag_data = []
        highlight_rows = set()
        
        for idx, tag in enumerate(resource_tags):
            tag_data.append([
                tag.subscription_name, tag.name, tag.value, tag.resource_name, 
                tag.resource_type, tag.resource_id, tag.resource_location,
                tag.compliance_status, tag.canonical_tag_name, tag.variation_matched
            ])
            if tag.is_mandatory_missing or tag.is_untagged:
                highlight_rows.add(idx + 2)
        
        # Tag values summary for resource tags
        tag_value_counts = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing:
                key = (tag.canonical_tag_name or tag.name, tag.value)
                tag_value_counts[key] = tag_value_counts.get(key, 0) + 1

        tag_value_summary = []
        for (name, value), count in sorted(tag_value_counts.items(), key=lambda x: x[1], reverse=True):
            tag_value_summary.append([name, value, count])

        self.create_enhanced_worksheet_with_table(
            wb, "Tag Values Summary", ["Tag Key", "Tag Value", "Usage Count"], tag_value_summary
        )
        self.create_enhanced_worksheet_with_table(
            wb, "Resource Tag Details", tag_headers, tag_data, 
            highlight_rows, compliance_columns=[8]
        )
        
        # Resource Group Tag Data
        rg_headers = ["Subscription", "RG Tag Name", "RG Tag Value", "Resource Group Name", 
                     "Resource Group ID", "Compliance Status", "Canonical Tag", "Variation Matched"]
        rg_data = []
        rg_highlight_rows = set()
        
        for idx, tag in enumerate(rg_tags):
            rg_data.append([
                tag.subscription_name, tag.name, tag.value, tag.resource_name, 
                tag.resource_id, tag.compliance_status, tag.canonical_tag_name, tag.variation_matched
            ])
            if tag.is_untagged:
                rg_highlight_rows.add(idx + 2)
        
        self.create_enhanced_worksheet_with_table(
            wb, "Resource Group Tag Details", rg_headers, rg_data, 
            rg_highlight_rows, compliance_columns=[6]
        )
    
    def generate_tag_summaries(self, wb: Workbook, resource_tags: List[TagData], 
                             rg_tags: List[ResourceGroupTagData]) -> None:
        """Generate tag summary sheets"""
        
        # Canonical tag summary
        canonical_counts = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing and tag.canonical_tag_name:
                canonical_counts[tag.canonical_tag_name] = canonical_counts.get(tag.canonical_tag_name, 0) + 1
        
        canonical_summary = [[name, count] for name, count in sorted(canonical_counts.items(), key=lambda x: x[1], reverse=True)]
        self.create_enhanced_worksheet_with_table(
            wb, "Canonical Tag Summary", ["Canonical Tag Name", "Usage Count"], canonical_summary
        )
        
        # Raw tag names summary
        tag_counts = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing:
                tag_counts[tag.name] = tag_counts.get(tag.name, 0) + 1
        
        tag_summary = [[name, count] for name, count in sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)]
        self.create_enhanced_worksheet_with_table(
            wb, "Raw Tag Names Summary", ["Tag Name", "Usage Count"], tag_summary
        )
        
        # Tag values summary
        tag_value_counts = {}
        for tag in resource_tags:
            if not tag.is_mandatory_missing:
                key = f"{tag.canonical_tag_name or tag.name}:{tag.value}"
                tag_value_counts[key] = tag_value_counts.get(key, 0) + 1
        
        tag_value_summary = []
        for key, count in sorted(tag_value_counts.items(), key=lambda x: x[1], reverse=True):
            name, value = key.split(":", 1)
            tag_value_summary.append([name, value, count])
        
        self.create_enhanced_worksheet_with_table